import os
import shutil
from datetime import datetime
import json
import uuid
from time import sleep

import openpyxl
import requests

from utils.PathManager import load_path_manager as lpm
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

file_master_data = lpm.input("Program Spec Master Data.xlsx")
file_fill_master_data = lpm.input("Program Spec Master Data.xlsx")
file_token = lpm.config("user_token.json")

'''
get user access token

JSON.parse(localStorage.getItem("persist:psg.auth")).refreshToken.replace(/"/g, "")
JSON.parse(localStorage.getItem("persist:psg.auth")).accessToken.replace(/"/g, "")

---------------------------------
get admin access token
localStorage.getItem("psg")

'''

REFRESH_TOKEN = None
BEARER_TOKEN = None

PORTAL_URL = ""
OCP_URL = ""


def refresh_token():
    url = ""

    global REFRESH_TOKEN
    payload = f'grant_type=refresh_token&refresh_token={REFRESH_TOKEN}&client_id=psg'
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9,zh-TW;q=0.8,zh;q=0.7',
        'Connection': 'keep-alive',
        'Content-type': 'application/x-www-form-urlencoded',
        'Origin': '',
        'Referer': '',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not/A)Brand";v="99", "Google Chrome";v="115", "Chromium";v="115"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"macOS"'
    }

    response = requests.request("POST", url, headers=headers, data=payload, verify=False)

    save_token_data(response.json())


def make_request(url, method='GET', params=None, data=None, headers=None, verify=False):
    if method == 'GET':
        response = requests.get(url, params=params, headers=headers, verify=verify)
    elif method == 'POST':
        response = requests.post(url, data=data, headers=headers, verify=verify)
    elif method == 'PUT':
        response = requests.put(url, data=data, headers=headers, verify=verify)
    else:
        raise ValueError("Unsupported HTTP method")

    return response


def load_sample_output_request_data():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["sample Output"]
    table_data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            table_data.append(row)

    wb.close()
    return table_data


def fill_sample_output_response_data(response_data):
    wb = openpyxl.load_workbook(file_master_data)
    sheet = wb["sample Output"]
    for item in response_data:
        ms_id, _, _, _, _, _, _, _, _, _ = item

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == ms_id:
                    print(f"Fill sample Output Data - {ms_id}")
                    sheet.cell(row=cell.row, column=4, value=item[3])
                    sheet.cell(row=cell.row, column=5, value=item[4])
                    sheet.cell(row=cell.row, column=6, value=item[5])
                    sheet.cell(row=cell.row, column=7, value=item[6])
                    sheet.cell(row=cell.row, column=8, value=item[7])
                    sheet.cell(row=cell.row, column=9, value=item[8])
                    sheet.cell(row=cell.row, column=10, value=item[9])

    wb.save(file_fill_master_data)


def save_token_data(data):
    with open(file_token, 'w') as file:
        json.dump(data, file)


def load_token_data():
    try:
        with open(file_token, 'r') as file:
            token_data = json.load(file)
            return token_data.get("refresh_token"), token_data.get("access_token")
    except FileNotFoundError:
        return None, None


def get_chrome_auto_download_token_file():
    download_dir = os.path.expanduser("~") + "/Downloads"

    source_file = os.path.join(download_dir, "user_token.json")

    if os.path.exists(source_file):

        target_file = file_token
        shutil.move(source_file, target_file)

        print(f"first run, success get chrome auto download token file:{target_file}")
    else:
        pass


def refresh_local_token_file():
    global REFRESH_TOKEN, BEARER_TOKEN
    REFRESH_TOKEN, BEARER_TOKEN = load_token_data()

    refresh_token()


def start():
    request_data = load_sample_output_request_data()
    response_data = []

    count = 0
    for item in request_data:
        if item[7] != "PASS":
            count += 1
            print(item[0])
    print(f"waiting execute {count}")

    # 1. make request get response
    for item in request_data:
        api_id = item[0]
        http_method = item[1]
        api_endpoint = item[2]
        path_parameters = item[3]
        request_body = item[4]
        run_it_flag = item[5]
        result = item[7]

        if run_it_flag == 'Y':

            if result != "PASS" or result is None:

                # compose path param to url
                if path_parameters is not None:
                    params = path_parameters.split(',')

                    for param in params:
                        key, value = param.split('=')
                        placeholder = f'{{{key}}}'
                        api_endpoint = api_endpoint.replace(placeholder, value)

                if "AA" in api_id:
                    request_url = OCP_URL + api_endpoint
                elif "BB" in api_id:
                    request_url = OCP_URL + api_endpoint
                elif "CC" in api_id and "admin" in api_endpoint:
                    request_url = OCP_URL + api_endpoint
                else:
                    request_url = PORTAL_URL + api_endpoint

                print("---------------------------------------------------")
                print(f"Process {api_id} - {request_url}")

                trace_id = str(uuid.uuid4())
                headers = {
                    'Content-Type': 'application/json; charset=UTF-8',
                    'X-Interaction-Id': trace_id,
                    'Authorization': f'Bearer {BEARER_TOKEN}',
                }

                response = make_request(request_url, http_method, None, request_body, headers)
                sleep(1)
                if response.status_code == 200 or response.status_code == 201:
                    response_body = json.dumps(response.json(), indent=4)

                    item = list(item)
                    item[9] = response_body
                    item[8] = trace_id
                    bizz_code = json.loads(response_body)["code"]
                    item[7] = "PASS" if (bizz_code == 200 or bizz_code == 201) else "FAIL"
                    item[6] = datetime.now().strftime('%Y-%m-%d')

                    item[5] = "N" if (bizz_code == 200 or bizz_code == 201) else "Y"  # finish test

                    print(f"    - {item[7]} - {item[9]}")
                    response_data.append(item)

                elif response.status_code == 400:
                    response_body = json.dumps(response.json(), indent=4)

                    item = list(item)
                    item[9] = response_body
                    item[8] = trace_id
                    item[7] = "FAIL"
                    item[6] = datetime.now().strftime('%Y-%m-%d')
                    item[5] = "Y"
                    print(f"    - {item[7]} - {item[9]}")
                    response_data.append(item)

                else:
                    # 404,401
                    print(response.status_code, response.content)

    # 2. fill response result data
    if response_data:
        fill_sample_output_response_data(response_data)

    print("finish")


if __name__ == '__main__':
    get_chrome_auto_download_token_file()
    refresh_local_token_file()
    start()
