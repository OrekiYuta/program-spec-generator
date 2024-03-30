import os
import shutil
import openpyxl
from utils.PathManager import load_path_manager as lpm

file_prog_spec_excel_template = lpm.template("prog_spec_excel_template.xlsx")
file_master_data = lpm.input("Program Spec Master Data.xlsx")

folder_input = str(lpm.input)
folder_transit_creator = str(lpm.transit.creator)


# Deprecated
def read_excel_section_deprecated(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    section_data = {}

    current_section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        section, api = row[:2]  # get A,B columns
        if section is not None:
            current_section = section
            section_data[current_section] = [api]
        elif api is not None and current_section is not None:
            section_data[current_section].append(api)

    wb.close()
    return section_data


def read_excel_section(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    section_data = {}

    section_col_idx = None
    api_col_idx = None
    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, col_name in enumerate(row, start=1):
            if col_name == "Module Section":
                section_col_idx = idx
            elif col_name == "API ID":
                api_col_idx = idx

    if section_col_idx is None or api_col_idx is None:
        raise ValueError("Columns 'Module Section' and 'API ID' not found in the Excel sheet.")

    for row in ws.iter_rows(min_row=2, values_only=True):
        module_section = row[section_col_idx - 1]
        api_id = row[api_col_idx - 1]

        if module_section is not None:
            current_section = module_section
            section_data[current_section] = [api_id]
        elif api_id is not None and current_section is not None:
            section_data[current_section].append(api_id)

    wb.close()
    return section_data


# Deprecated
def copy_worksheets(source_wb, target_wb):
    for sheet in source_wb.sheetnames:
        source_ws = source_wb[sheet]
        target_ws = target_wb.create_sheet(title=sheet)

        for row in source_ws.iter_rows(values_only=True):
            target_ws.append(row)


def create_folders_and_files(table_data, output_folder):
    print("create folders and files - start")
    for folder, files in table_data.items():

        safe_folder_name = folder.replace("/", "_")
        folder_path = os.path.join(output_folder, safe_folder_name)
        os.makedirs(folder_path, exist_ok=True)
        print(safe_folder_name)

        for file in files:
            # handle special character in file name e.g. "/"
            # safe_file_name = file.replace("/", "_")
            # NOTE UserWarning: Title is more than 31 characters. Some applications may not be able to read the file

            excel_path = os.path.join(folder_path, f"{file}.xlsx")
            shutil.copy2(file_prog_spec_excel_template, excel_path)
            print("-   " + file + ".xlxs")

    print("create folders and files - end")


def start():
    print(">>>>>>>>>>> CREATOR PROCESS START >>>>>>>>>>>")
    table_data = read_excel_section(file_master_data, "MASTER")
    create_folders_and_files(table_data, folder_transit_creator)
    print(">>>>>>>>>>> CREATOR PROCESS END >>>>>>>>>>>>>")


if __name__ == '__main__':
    start()
