import os

import openpyxl
import yaml

from utils.PathManager import load_path_manager as lpm

file_master_data = lpm.input("Program Spec Master Data.xlsx")
file_db_schema = lpm.input("DB-Schema.xlsx")
file_access_control = lpm.input("Access Control.xlsx")
file_validation_rules = lpm.input("Validation Rules.xlsx")

folder_swagger = str(lpm.input.Swagger)
folder_transit_creator = str(lpm.transit.creator)

result = []

def load_swagger_file_data(swagger_file):
    with open(swagger_file, 'r') as file:
        swagger_data = yaml.safe_load(file)

    return swagger_data


def process_components_schemas(components_schemas):
    new_schemas = {}
    for schema_name, schema_data in components_schemas.items():
        if 'properties' in schema_data:
            properties = schema_data['properties']
            for prop_name, prop_data in properties.items():
                for i_prop_name, i_prop_data in prop_data.items():
                    if isinstance(i_prop_data, dict) and '$ref' in i_prop_data:
                        ref = i_prop_data['$ref']
                        ref_name = ref.split('/')[-1]
                        if ref_name in components_schemas:
                            new_properties = components_schemas[ref_name]
                            # append $ref mapping item to the current node
                            i_prop_data[ref_name] = new_properties

        new_schemas[schema_name] = schema_data

    return new_schemas


def find_ref(paths_schemas, components_schemas):
    ref = []

    if isinstance(paths_schemas, dict):
        if "$ref" in paths_schemas:
            ref.append(paths_schemas["$ref"])
            ref_name = ref[0].split('/')[-1]

            if ref_name in components_schemas:
                new_properties = components_schemas[ref_name]
                # paths_schemas["$ref"] = new_properties
                paths_schemas[ref_name] = new_properties

        else:
            for value in paths_schemas.values():
                ref.extend(find_ref(value, components_schemas))
    elif isinstance(paths_schemas, list):
        for item in paths_schemas:
            ref.extend(find_ref(item, components_schemas))

    return ref


def process_paths_schemas(paths_schemas, components_schemas):
    find_ref(paths_schemas, components_schemas)
    return paths_schemas


def extract_data_from_swagger(swagger_file):
    data_swag = load_swagger_file_data(swagger_file)

    processed_components_schemas = process_components_schemas(data_swag.get('components')['schemas'])

    compose_data_swag = process_paths_schemas(data_swag.get('paths'), processed_components_schemas)

    return compose_data_swag


def clear_extract_swagger_data(extract_swagger_data):
    # print(swagger_data)
    clear_swagger_data_list = []
    clear_swagger_data_item = {
        "API End Point": None,
        "HTTP Method": None,

        # Input parameters
        # - Path Parameter(s)
        "IP-PP": None,
        # Input parameters
        # - Request Body
        "IP-RB": None,
        # Business Rules & Logic (Description / Pseudo Code)
        # - Data Access Layer
        #   - Operation: Read / Write
        "BRL-DAL": None
    }

    for api_end_point, http_method in extract_swagger_data.items():
        for method_name, inner_data in http_method.items():

            current_item = dict(clear_swagger_data_item)
            current_item["API End Point"] = api_end_point
            current_item["HTTP Method"] = method_name

            if inner_data.get("parameters") is not None:
                current_item["IP-PP"] = inner_data.get("parameters")

            # if inner_data.get("requestBody") is not None:
            if (
                    inner_data.get("requestBody")
                    and inner_data["requestBody"].get("content")
                    and inner_data["requestBody"]["content"].get("application/json")
                    and inner_data["requestBody"]["content"]["application/json"].get("schema")
            ):
                # current_item["IP-RB"] = inner_data.get("requestBody")
                current_item["IP-RB"] = inner_data["requestBody"]["content"]["application/json"].get("schema")

            # print(api_end_point + "  " + method_name)
            if (
                    inner_data.get("responses")
                    and inner_data["responses"].get("200")
                    and inner_data["responses"]["200"].get("content")
                    and inner_data["responses"]["200"]["content"].get("*/*")
                    and inner_data["responses"]["200"]["content"]["*/*"].get("schema")
            ):
                current_item["BRL-DAL"] = inner_data["responses"]["200"]["content"]["*/*"]["schema"]

            clear_swagger_data_list.append(current_item)

    return clear_swagger_data_list


def extract_data_from_master_data_excel():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["MASTER"]

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    data_list = []

    current_module_section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_map = {}
        for col_idx, value in enumerate(row, start=1):

            col_name = header[col_idx - 1]

            # assort module section
            if col_idx == 1:
                if value is not None:
                    current_module_section = value
                else:
                    value = current_module_section

            data_map[col_name] = value

        data_list.append(data_map)

    wb.close()
    # print(data_list)
    return data_list


def load_excel_table_data_rows(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    table_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            table_data.append(row)
    wb.close()
    return table_data


def fill_ip_pp_table_default_value(fill_data_array):
    default_value_data = load_excel_table_data_rows(file_master_data, "IP-PP-Default-Value")

    for ip_pp_item in fill_data_array:
        parameters_name = ip_pp_item['Parameters Name']
        for default_item in default_value_data:
            if default_item[0] == parameters_name:
                ip_pp_item['Description'] = default_item[1]
                break

    return fill_data_array


def fill_ip_pp_table_data(wb, inc_data):
    print("     -> Fill Input parameters - Path Parameter(s)")
    ip_pp_table = []
    ip_pp_table_cow = {
        "Parameters Name": "<checkme>",
        "Description": "<checkme>",
        "Mandatory": "<checkme>",
    }

    ws = wb["IP-PP-TABLE"]

    for data in inc_data:

        if "path" == data["in"]:
            # e.g. [{'name': 'requestUuid', 'in': 'path', 'required': True, 'schema': {'type': 'string', 'format': 'uuid'}}]
            current_ip_pp_table_cow = ip_pp_table_cow.copy()
            current_ip_pp_table_cow["Parameters Name"] = data["name"]
            current_ip_pp_table_cow["Mandatory"] = "Y" if data["required"] else "N"

            ip_pp_table.append(current_ip_pp_table_cow)
        else:
            # e.g. [{'name': 'Authorization', 'in': 'header', 'required': True, 'schema': {'type': 'string'}}]
            print(f'     !!! warning param is not in path, is in {data["in"]}')
            pass

    # fill default value [Description]
    final_array = fill_ip_pp_table_default_value(ip_pp_table)

    remove_all_rows_except_first(ws)
    # set data in sheet
    for row_data in final_array:
        row = [row_data["Parameters Name"],
               row_data["Description"],
               row_data["Mandatory"]]
        ws.append(row)


def fill_ip_rb_table_default_value(fill_data_array):
    default_value_data = load_excel_table_data_rows(file_master_data, "IP-RB-Default-Value")

    for rb_item in fill_data_array:
        parameters_name = rb_item['Parameters Name']

        no_matched_flag = False
        for default_item in default_value_data:
            if default_item[0] == parameters_name:
                rb_item['Possible Values'] = default_item[1]
                rb_item['Description'] = default_item[2]
                no_matched_flag = True
                break

        if no_matched_flag is False:
            print(f"         !!! warning field [Parameters Name]=<{parameters_name}> no_matched, "
                  f"pls check Master data sheet IP-RB-Default-Value")

    return fill_data_array


def fill_ip_rb_table_data(wb, inc_data):
    print("     -> Fill Input parameters - Request Body")

    ws = wb["IP-RB-TABLE"]

    keys = list(inc_data.keys())  # get all key name
    data = inc_data[keys[1]]  # get second item

    required_array = []  # Mandatory key
    if data.get("required"):
        required_array = data["required"]

    parameters_array = compose_request_body_parameters(data["properties"])

    # set Mandatory value
    for required_item in required_array:
        for param in parameters_array:
            if required_item in param["Parameters Name"]:
                param["Mandatory"] = "Y"
                break

    # fill default value [Description]
    final_array = fill_ip_rb_table_default_value(parameters_array)

    remove_all_rows_except_first(ws)
    # set data in sheet
    for row_data in final_array:
        row = [
            row_data["Parameters Name"],
            row_data["Possible Values"],
            row_data["Description"],
            row_data["Mandatory"]
        ]
        ws.append(row)


def fill_swagger_data_to_each_dist(swg_data):
    print("------> FILL SWAGGER DATA START")

    wb = openpyxl.load_workbook(file_master_data)
    sheet = wb["MASTER"]

    excel_headers = [cell.value for cell in sheet[1]]

    api_endpoint_column_index = excel_headers.index("API End Point")
    http_method_column_index = excel_headers.index("HTTP Method")
    module_section_column_index = excel_headers.index("Module Section")
    api_id_column_index = excel_headers.index("API ID")
    operation_flag_index = excel_headers.index("Data Access Layer Operation")
    read_operation_table_index = excel_headers.index("Database Operation Table - Read")
    write_operation_table_index = excel_headers.index("Database Operation Table - Write")

    for swg_item_data in swg_data:
        api_endpoint = swg_item_data["API End Point"]
        http_method = swg_item_data["HTTP Method"]

        matching_rows = []

        current_module_section = None

        for row in sheet.iter_rows(min_row=2, values_only=True):

            # handle module section
            if row[module_section_column_index] is not None:
                current_module_section = row[module_section_column_index]

            # print("search:  " + http_method + "   " + api_endpoint.replace("/internal", ""))
            # print("target:  " + row[http_method_column_index] + "   " + row[api_endpoint_column_index])

            if row[http_method_column_index].lower().strip() == http_method.lower().strip():
                if (
                        api_endpoint.lower().strip().replace("/internal", "") ==
                        row[api_endpoint_column_index].lower().strip().replace(" ", "")
                                .replace("/cas/psg/v1", "")  # common swg
                                .replace("/cas/psg/batch/v1/internal", "")  # batch swg
                                .replace("/cas/psg/v1", "")  # cvh swg

                ):
                    row_list = list(row)
                    row_list[module_section_column_index] = current_module_section  # handle module section
                    matching_rows.append(tuple(row_list))

                    # print("=> ping:    " + row[api_id_column_index] + "   "
                    #       + row[http_method_column_index] + "   "
                    #       + row[api_endpoint_column_index])
                    break

        if matching_rows:
            module_section = matching_rows[0][module_section_column_index]
            api_id = matching_rows[0][api_id_column_index]
            operation_flag = matching_rows[0][operation_flag_index]

            read_operation_table = matching_rows[0][read_operation_table_index]
            write_operation_table = matching_rows[0][write_operation_table_index]

            module_section = module_section.replace("/", "_")
            module_folder_path = os.path.join(folder_transit_creator, module_section)
            xlsx_file_path = os.path.join(module_folder_path, f"{api_id}.xlsx")

            print("=> " + module_section + " => " + api_id + " -> "
                  + "[ " + http_method + "  " + api_endpoint + " ]")

            if os.path.exists(xlsx_file_path):
                target_api_wb = openpyxl.load_workbook(xlsx_file_path)

                if swg_item_data.get("IP-PP") is not None:
                    fill_ip_pp_table_data(target_api_wb, swg_item_data.get("IP-PP"))

                if swg_item_data.get("IP-RB") is not None:
                    fill_ip_rb_table_data(target_api_wb, swg_item_data.get("IP-RB"))

                if swg_item_data.get("BRL-DAL") is not None:

                    if operation_flag is not None:
                        '''
                            1.Read use DRL-DAL(API responses) define table row, then mapping DB Schema
                            2.Write use DB Schema define table row, 
                                 then mapping: 
                                    - IP-IP(API path parameter(s))
                                    - IP-RB(API requestBody) 
                        '''
                        if "Read" in operation_flag:
                            fill_brl_dal_r_table_data(target_api_wb, swg_item_data.get("BRL-DAL"),
                                                      read_operation_table)

                        if "Write" in operation_flag:
                            '''
                            1.have path param, no requestBody 
                            2.both have path param and requestBody
                            3.no path param , have requestBody

                            '''
                            # print(f"write - ip-pp {swg_item_data.get('IP-PP')}")
                            # print(f"write - ip-rb {swg_item_data.get('IP-RB')}")

                            fill_brl_dal_w_table_data(target_api_wb,
                                                      swg_item_data.get('HTTP Method'),
                                                      api_endpoint,
                                                      swg_item_data.get("IP-PP"),
                                                      swg_item_data.get("IP-RB"),
                                                      write_operation_table)

                target_api_wb.save(xlsx_file_path)

            else:
                print("   !!! warning missing " + module_section + " - " + api_id + " - "
                      + "[ " + http_method + "  " + api_endpoint + " ]")

    wb.close()

    print("------> FILL SWAGGER DATA END")


def compose_ip_rb_fill_brl_dal_w_table(inc_ip_rb_data):
    keys = list(inc_ip_rb_data.keys())
    data = inc_ip_rb_data[keys[1]]
    parameters_array = []

    if isinstance(data, dict):

        properties_flag = data.get("properties")
        if properties_flag is None:
            '''
              responses:
                "400":
                  description: Bad Request
                "200":
                  description: OK
                  content:
                    '*/*':
                      schema:
                        type: object
                        additionalProperties:
                          type: object
            '''
            print("         !!! warning the api response body is a map, "
                  "rather than an entity, can not retrieve properties")

            param = {
                "Source": "(pls update swagger)",
                "Source Field Name": "(pls update swagger)",
                "Destination": "(pls update swagger)",
                "Destination Field Name": "(pls update swagger)",
                "Conversion Logic": "(pls update swagger)"
            }
            parameters_array.append(param)

        else:
            # normal scenario
            parameters_array = compose_write_response_body_parameters(data["properties"])

    else:
        '''
        responses:
          "400":
            description: Bad Request
          "200":
            description: OK
            content:
              '*/*':
                schema:
                  type: string
                  format: uuid
        '''
        print("         !!! warning the api response body is a string, also is normal scenario"
              "(using entity and update swagger will great handle this)")

        param = {
            "Source": "Request Body",
            "Source Field Name": "recUuid",  # Note here is Hardcode
            "Destination": "<checkme>",
            "Destination Field Name": "<checkme>",
            "Conversion Logic": "<checkme>"
        }
        parameters_array.append(param)

    return parameters_array


def compose_ip_pp_fill_brl_dal_w_table(inc_ip_pp_data):
    parameters_array = []
    # [{'name': 'requestUuid', 'in': 'path', 'required': True, 'schema': {'type': 'string', 'format': 'uuid'}}]
    for data in inc_ip_pp_data:
        if "path" == data["in"]:

            param = {
                "Source": "Path Parameter(s)",
                "Source Field Name": data["name"],
                "Destination": "<checkme>",
                "Destination Field Name": "<checkme>",
                "Conversion Logic": "<checkme>"
            }
            parameters_array.append(param)
        else:
            print(f'     !!! warning param is not in path, is in {data["in"]}')
            pass

    return parameters_array


def fill_brl_dal_w_table_default_value(fill_data_array):
    default_value_data = load_excel_table_data_rows(file_master_data, "BRL-DAL-W-Default-Value")

    for result_item in fill_data_array:

        source_field_name = result_item['Source Field Name']
        destination = result_item['Destination']

        matched_destination_array = []
        matched_flag = False
        for default_item in default_value_data:
            # if default_item[0].strip() == source_field_name:
            if default_item[1].upper().strip() == destination:
                matched_flag = True
                matched_destination_array.append(default_item)

        if matched_flag:

            if len(matched_destination_array) > 1:
                matched_item = None

                for temp_item in matched_destination_array:
                    # if temp_item[1].upper().strip() == destination:
                    if temp_item[0].strip() == source_field_name:
                        matched_item = temp_item
                        break

                if matched_item is not None:
                    result_logic = matched_item[2]
                    result_item['Conversion Logic'] = result_logic
                else:
                    print(f"         !!! warning field [Destination]=<{destination}> matched "
                          f"but [Source Field Name]=<{source_field_name}> no_matched, "
                          f"pls check Master data sheet BRL-DAL-W-Default-Value")
                    result_logic = matched_destination_array[0][2]
                    result_item['Conversion Logic'] = result_logic
            else:
                result_logic = matched_destination_array[0][2]
                result_item['Conversion Logic'] = result_logic

        else:
            print(f"         !!! warning field [Destination]=<{destination}>no_matched, "
                  f"pls check Master data sheet BRL-DAL-W-Default-Value")

    return fill_data_array


def fill_brl_dal_w_table_data(wb, http_method, endpoint, inc_ip_pp_data, inc_ip_rb_data, db_table):
    print("     -> Fill Business Rules & Logic (Description / Pseudo Code) - Data Access Layer - Operation: Write")

    ws = wb["BRL-DAL-W-TABLE"]

    pp_parameters_array = []
    rb_parameters_array = []
    merged_parameters_array = []

    if inc_ip_pp_data:
        # handle path parameter(s)
        pp_parameters_array = compose_ip_pp_fill_brl_dal_w_table(inc_ip_pp_data)

    if inc_ip_rb_data:
        # handle requestBody
        rb_parameters_array = compose_ip_rb_fill_brl_dal_w_table(inc_ip_rb_data)

    merged_parameters_array.extend(pp_parameters_array)
    merged_parameters_array.extend(rb_parameters_array)

    # extract db schema
    db_schema_map = extract_db_schema(db_table)

    # loop db db_schema_map ,then mapping parameters_array
    result_array = fill_write_mapping_db_schema_value(merged_parameters_array, db_schema_map)

    # fill default value [Conversion Logic]
    final_array = fill_brl_dal_w_table_default_value(result_array)

    remove_all_rows_except_first(ws)
    # set data in sheet
    for row_data in final_array:
        if http_method == 'put' and endpoint.endswith("download-status"):
            if row_data["Source Field Name"] == "fileDownloadStatus":
                row = [
                    row_data["Source"],
                    row_data["Source Field Name"],
                    row_data["Destination"],
                    row_data["Destination Field Name"],
                    'Fixed value: "Y"'
                ]
                ws.append(row)
            elif row_data["Source Field Name"] == "fileLastDownloadUser":
                row = [
                    row_data["Source"],
                    row_data["Source Field Name"],
                    row_data["Destination"],
                    row_data["Destination Field Name"],
                    'The user name triggered the request'
                ]
                ws.append(row)
            elif row_data["Source Field Name"] == "fileLastDownloadTime":
                row = [
                    row_data["Source"],
                    row_data["Source Field Name"],
                    row_data["Destination"],
                    row_data["Destination Field Name"],
                    'Current system time'
                ]
                ws.append(row)
        elif http_method == 'put' and row_data["Source"] == "N/A" \
                and row_data["Destination"] not in ["CMN_PORTAL_DOC_REF", "CMN_PORTAL_CONVSN"]\
                and "direct mapping" in row_data["Conversion Logic"].lower():
            row = [
                row_data["Source"],
                row_data["Source Field Name"],
                row_data["Destination"],
                row_data["Destination Field Name"],
                "Persist the existing value"
            ]
            ws.append(row)
        else:
            row = [
                row_data["Source"],
                row_data["Source Field Name"],
                row_data["Destination"],
                row_data["Destination Field Name"],
                row_data["Conversion Logic"]
            ]
            ws.append(row)

            if row_data["Conversion Logic"] == "<checkme>":
                global result
                result.append(f'{row_data["Source"]}	{row_data["Source Field Name"]}	{row_data["Destination"]}	{row_data["Destination Field Name"]}	{row_data["Conversion Logic"]}')


def fill_write_mapping_db_schema_value(parameters_array, db_schema_map):
    write_parameters_array = []
    for table_name, schema_list in db_schema_map.items():
        for schema_item in schema_list:
            param = {
                "Source": "<checkme>",
                "Source Field Name": "<checkme>",
                "Destination": table_name,
                "Destination Field Name": schema_item["COLUMN_NAME"],
                "Conversion Logic": "<checkme>",
                "ACTUAL API RESPONSE NAME": schema_item["ACTUAL API RESPONSE NAME"],  # using for data compose
                "PROPERTY": schema_item["PROPERTY"]  # using for data compose
            }
            write_parameters_array.append(param)

    matched_flag = False
    for write_params in write_parameters_array:
        for params in parameters_array:
            if params["Source Field Name"] == write_params["PROPERTY"]:
                # second alternate mapping
                matched_flag = True
                write_params["Source"] = params["Source"]
                write_params["Source Field Name"] = params["Source Field Name"]
                break

        if matched_flag is False:
            # print(params["Source Field Name"] + "-> no matched")
            pass

    for write_params in write_parameters_array:
        # third alternate mapping
        if write_params["Source Field Name"] == "<checkme>":
            write_params["Source"] = "N/A"
            write_params["Source Field Name"] = write_params["PROPERTY"]

    # remove write_parameters_array key "ACTUAL API RESPONSE NAME","PROPERTY"
    for item in write_parameters_array:
        del item["ACTUAL API RESPONSE NAME"]
        del item["PROPERTY"]

    return write_parameters_array


def extract_db_schema(table_name_str):
    extracted_data = {}
    if table_name_str is not None:
        table_names = table_name_str.strip().replace(" ", "").split(",")
        workbook = openpyxl.load_workbook(file_db_schema)

        for table_name in table_names:
            data_list = []
            if table_name in workbook.sheetnames:
                sheet = workbook[table_name]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data_item = {
                        "COLUMN_NAME": row[0],
                        "DATA_TYPE": row[1],
                        "PROPERTY": row[2],
                        "ACTUAL API RESPONSE NAME": row[3],
                    }
                    data_list.append(data_item)

            extracted_data[table_name] = data_list

        workbook.close()

        # print(extracted_data)
    return clear_db_schema(extracted_data)


def clear_db_schema(db_schema_data):
    for table_name, items in db_schema_data.items():
        cleaned_items = []
        for i_item in items:

            values = list(i_item.values())
            if all(value is None for value in values):
                continue

            cleaned_items.append(i_item)

        db_schema_data[table_name] = cleaned_items

    return db_schema_data


def fill_read_mapping_db_schema_value(parameters_array, db_schema_map):
    for parameters in parameters_array:
        destination_field_name = parameters["Destination Field Name"]

        matched_flag = False
        for table_name, schema_list in db_schema_map.items():
            for schema_item in schema_list:
                if schema_item["ACTUAL API RESPONSE NAME"] == destination_field_name:
                    # print(destination_field_name + "-> matching ->" + schema_item["ACTUAL API RESPONSE NAME"])
                    matched_flag = True
                    parameters["Source Table"] = table_name
                    parameters["Source Field Name"] = schema_item["COLUMN_NAME"]
                    break

                elif schema_item["PROPERTY"] == destination_field_name:
                    # second alternate mapping
                    parameters["Source Table"] = table_name
                    parameters["Source Field Name"] = schema_item["COLUMN_NAME"]
                    break

                elif destination_field_name.lower() in schema_item["PROPERTY"].lower():
                    # third alternate mapping
                    parameters["Source Table"] = table_name
                    parameters["Source Field Name"] = schema_item["COLUMN_NAME"]
                    break

                else:
                    pass

        if matched_flag is False:
            print("         !!! warning " + destination_field_name +
                  " no mapping in " + table_name + " set N/A ")
            # print(parameters["Source Table"] + " " + parameters["Source Field Name"])
            if parameters["Source Table"] == "<checkme>":
                parameters["Source Table"] = "N/A"
            if parameters["Source Field Name"] == "<checkme>":
                parameters["Source Field Name"] = "N/A"

    return parameters_array


def fill_brl_dal_r_table_default_value(fill_data_array):
    default_value_data = load_excel_table_data_rows(file_master_data, "BRL-DAL-R-Default-Value")

    for result_item in fill_data_array:

        source_table = result_item['Source Table']
        destination_field_name = result_item['Destination Field Name']

        matched_destination_field_name_array = []
        matched_flag = False
        for default_item in default_value_data:
            if default_item[1].strip() == destination_field_name:
                matched_flag = True
                matched_destination_field_name_array.append(default_item)

        if matched_flag:

            if len(matched_destination_field_name_array) > 1:
                matched_item = None

                for temp_item in matched_destination_field_name_array:
                    if temp_item[0].upper().strip() == source_table:
                        matched_item = temp_item
                        break

                if matched_item is not None:
                    result_logic = matched_item[2]
                    result_item['Conversion Logic'] = result_logic
                else:
                    print(f"         !!! warning field [Destination Field Name]=<{destination_field_name}> matched "
                          f"but [Source Table]=<{source_table}> no_matched, "
                          f"pls check Master data sheet BRL-DAL-R-Default-Value")
                    result_logic = matched_destination_field_name_array[0][2]
                    result_item['Conversion Logic'] = result_logic

            else:
                result_logic = matched_destination_field_name_array[0][2]
                result_item['Conversion Logic'] = result_logic

        else:
            print(f"         !!! warning field [Destination Field Name]=<{destination_field_name}> no_matched, "
                  f"pls check Master data sheet BRL-DAL-R-Default-Value")

    return fill_data_array


def fill_brl_dal_r_table_data(wb, inc_data, db_table):
    print("     -> Fill Business Rules & Logic (Description / Pseudo Code) - Data Access Layer - Operation: Read")

    ws = wb["BRL-DAL-R-TABLE"]

    keys = list(inc_data.keys())
    data = inc_data[keys[1]]
    parameters_array = []

    if isinstance(data, dict):
        properties_flag = data.get("properties")

        if properties_flag is None:
            '''
              responses:
                "400":
                  description: Bad Request
                "200":
                  description: OK
                  content:
                    '*/*':
                      schema:
                        type: object
                        additionalProperties:
                          type: object
            '''
            print("         !!! warning the api response body is a map, "
                  "rather than an entity, can not retrieve properties")

            param = {
                "Source Table": "(pls update swagger)",
                "Source Field Name": "(pls update swagger)",
                "Destination Entity Name": "(pls update swagger)",
                "Destination Field Name": "(pls update swagger)",
                "Conversion Logic": "(pls update swagger)"
            }
            parameters_array.append(param)

        else:
            # normal scenario
            parameters_array = compose_read_response_body_parameters(data["properties"])

    else:
        '''
        responses:
          "400":
            description: Bad Request
          "200":
            description: OK
            content:
              '*/*':
                schema:
                  type: string
                  format: uuid
        '''
        print("         --- the api response body is a string, also is normal scenario, "
              "updated to pojo is best practice")

        param = {
            "Source Table": "<checkme>",
            "Source Field Name": "<checkme>",
            "Destination Entity Name": "Response.payload",
            "Destination Field Name": "recUuid",  # Note here is Hardcode
            "Conversion Logic": "<checkme>"
        }
        parameters_array.append(param)

    # extract db schema
    db_schema_map = extract_db_schema(db_table)

    # loop parameters_array ,then mapping db schema
    result_array = fill_read_mapping_db_schema_value(parameters_array, db_schema_map)

    # fill default value [Conversion Logic]
    final_array = fill_brl_dal_r_table_default_value(result_array)

    remove_all_rows_except_first(ws)
    # set data in sheet
    for row_data in final_array:
        row = [
            row_data["Source Table"],
            row_data["Source Field Name"],
            row_data["Destination Entity Name"],
            row_data["Destination Field Name"],
            row_data["Conversion Logic"]
        ]
        ws.append(row)


def compose_request_body_parameters(schema, parent_key=""):
    parameters = []

    for key, value in schema.items():
        current_key = key if not parent_key else f"{parent_key}.{key}"

        param = {
            "Parameters Name": current_key,
            "Possible Values": "<checkme>",
            "Description": "<checkme>",
            "Mandatory": "N"
        }

        parameters.append(param)

        if "items" in value and value["type"] == "array":
            # print("obj  - " + str(value))
            item_obj = value["items"]
            # print(item_obj.keys())
            keys = list(item_obj.keys())

            last_item_obj = item_obj[keys[-1]]  # get the last item obj
            if isinstance(last_item_obj, dict):
                # print("obj.last  - " + str(last_item_obj))
                array_item_props = last_item_obj.get("properties")
                if array_item_props is not None:

                    array_params = compose_request_body_parameters(
                        array_item_props,
                        current_key + "[]"
                    )
                    parameters.extend(array_params)
                else:
                    pass
                    # print("obj-else-v  - " + str(value))
                    # print("obj-else-k  - " + str(keys))
            else:
                pass
                # print("obj.last-not-is-dict  - " + str(last_item_obj))

    return parameters


def compose_read_response_body_parameters(schema, parent_key="", n_a_set_only_once_flag=True):
    parameters = []

    for key, value in schema.items():
        current_key = key if not parent_key else f"{parent_key}.{key}"

        # handle the split string display
        if "." in current_key:
            parts = current_key.split(".")
            den_str = parts[0]
            dfn_str = parts[-1]
        else:
            den_str = None
            dfn_str = current_key

        param = {
            "Source Table": "<checkme>",
            "Source Field Name": "<checkme>",
            "Destination Entity Name": "Response.payload" + ("." + den_str if den_str else ""),
            "Destination Field Name": dfn_str,
            "Conversion Logic": "<checkme>"
        }

        parameters.append(param)

        if "items" in value and value["type"] == "array":

            if n_a_set_only_once_flag:
                '''
                   N/A  Set only once, avoid multiple nesting items array
                   e.g. 
                       schema = [
                           totalCount:[] 
                           result:[
                               ...
                               items:[ <- Note here
                                   ...
                                   properties:[
                                       trCode:[]
                                       documents:[
                                           items:[ <- Note here
                                           ]
                                       ]
                                       conversations:[
                                            items:[ <- Note here
                                           ]
                                       ]
                                   ]
                               ]
                           ]
                       ]
                '''
                n_a_set_only_once_flag = False
                # Note handle N/A for payload array [Source Table],[Source Field Name]
                # set the Node field to N/A
                '''
                    Source Table 	Source Field Name 	Destination Entity Name 	Destination Field Name 
                    N/A	            N/A	                Response.payload	        totalCount
                    N/A	            N/A	                Response.payload	        results   
                '''
                for item in parameters:
                    item['Source Table'] = 'N/A'
                    item['Source Field Name'] = 'N/A'
            else:
                # set the current field to N/A
                for item in parameters:
                    if item["Destination Field Name"] == key:
                        item['Source Table'] = 'N/A'
                        item['Source Field Name'] = 'N/A'

            # print("obj  - " + str(value))
            item_obj = value["items"]
            # print(item_obj.keys())
            keys = list(item_obj.keys())

            last_item_obj = item_obj[keys[-1]]  # get the last item obj
            if isinstance(last_item_obj, dict):
                # print("obj.last  - " + str(last_item_obj))
                array_item_props = last_item_obj.get("properties")
                if array_item_props is not None:

                    array_params = compose_read_response_body_parameters(
                        array_item_props,
                        current_key + "[]",
                        n_a_set_only_once_flag
                    )
                    parameters.extend(array_params)
                else:
                    pass
                    # print("obj-else-v  - " + str(value))
                    # print("obj-else-k  - " + str(keys))
            else:
                pass
                # print("obj.last-not-is-dict  - " + str(last_item_obj))

    return parameters


def compose_write_response_body_parameters(schema, parent_key=""):
    parameters = []

    for key, value in schema.items():
        current_key = key if not parent_key else f"{parent_key}.{key}"

        # handle the split string display
        if "." in current_key:
            parts = current_key.split(".")
            den_str = "." + parts[0]
            dfn_str = parts[-1]
        else:
            den_str = None
            dfn_str = current_key

        param = {
            "Source": "RequestBody" + (den_str if den_str else ""),
            "Source Field Name": dfn_str,
            "Destination": "<checkme>",
            "Destination Field Name": "<checkme>",
            "Conversion Logic": "<checkme>"
        }

        parameters.append(param)

        if "items" in value and value["type"] == "array":
            # print("obj  - " + str(value))
            item_obj = value["items"]
            # print(item_obj.keys())
            keys = list(item_obj.keys())

            last_item_obj = item_obj[keys[-1]]  # get the last item obj
            if isinstance(last_item_obj, dict):
                # print("obj.last  - " + str(last_item_obj))
                array_item_props = last_item_obj.get("properties")
                if array_item_props is not None:

                    array_params = compose_write_response_body_parameters(
                        array_item_props,
                        current_key + "[]"
                    )
                    parameters.extend(array_params)
                else:
                    pass
                    # print("obj-else-v  - " + str(value))
                    # print("obj-else-k  - " + str(keys))
            else:
                pass
                # print("obj.last-not-is-dict  - " + str(last_item_obj))

    return parameters


def remove_all_rows_except_first(ws):
    rows_to_delete = ws.max_row - 1
    if rows_to_delete > 0:
        ws.delete_rows(2, amount=rows_to_delete)


def fill_master_data_to_each_dist(master_data):
    print("------> FILL MASTER DATA START")

    for api_item in master_data:
        module_section = api_item.get("Module Section")
        api_id = api_item.get("API ID")

        # handle special character in folder name e.g. "/"
        module_section = module_section.replace("/", "_")

        module_folder_path = os.path.join(folder_transit_creator, module_section)

        xlsx_file_path = os.path.join(module_folder_path, f"{api_id}.xlsx")
        print("=> " + module_section + " => " + api_id)
        if os.path.exists(xlsx_file_path):

            wb = openpyxl.load_workbook(xlsx_file_path)
            ws = wb["MASTER"]

            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            data_to_fill = []
            for key, value in api_item.items():
                if key in header_row:
                    col_idx = header_row.index(key) + 1
                    data_to_fill.append((col_idx, value))

            # new_row_idx = ws.max_row + 1
            new_row_idx = 2  # set data in second row
            # if the second row has data, delete it
            if ws.cell(row=new_row_idx, column=1).value is not None:
                ws.delete_rows(new_row_idx)
            # set the new data
            for col_idx, value in data_to_fill:
                ws.cell(row=new_row_idx, column=col_idx, value=value)

            wb.save(xlsx_file_path)

    print("------> FILL MASTER DATA END")


def extract_data_from_access_control_excel():
    return extract_header_table_data_from_excel(file_access_control)


def extract_header_table_data_from_excel(file_excel):
    workbook = openpyxl.load_workbook(file_excel)

    result_data = []

    for sheet_name in workbook.sheetnames:
        sheet_data = []
        sheet = workbook[sheet_name]

        if sheet.title != "README":

            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data_dict = {}
                for header, value in zip(headers, row):
                    data_dict[header] = value
                sheet_data.append(data_dict)

            result_data.append({sheet_name: sheet_data})

    workbook.close()

    # print(result_data)
    return result_data


def compose_module_section_with_api_subset_data(subset_array):
    master_data = extract_data_from_master_data_excel()

    compose_data_array = []

    for subset_item in subset_array:
        for subset_key, subset_value in subset_item.items():
            for master_item in master_data:
                if subset_key == master_item['API ID']:
                    compose_data_array.append({
                        'Module Section': master_item['Module Section'],
                        'API ID': subset_key,
                        'Sub Set': subset_value
                    })

    # print(compose_data_array)
    return compose_data_array


def fill_access_control_data_to_each_dist(ac_data_array):
    print("------> FILL ACCESS CONTROL DATA START")

    compose_data = compose_module_section_with_api_subset_data(ac_data_array)

    for item in compose_data:
        module_section = item["Module Section"]
        safe_name_module_section = module_section.replace("/", "_")  # handle special character in folder name e.g. "/"
        api_id = item["API ID"]
        sub_set_data = item["Sub Set"]

        module_folder_path = os.path.join(folder_transit_creator, safe_name_module_section)
        xlsx_file_path = os.path.join(module_folder_path, f"{api_id}.xlsx")

        if os.path.exists(xlsx_file_path):
            target_api_wb = openpyxl.load_workbook(xlsx_file_path)

            ws = target_api_wb["BRL-AC-TABLE"]

            if not sub_set_data:
                print(f"   -> Fill Access Control / {api_id} No Access Control")
            else:
                print(f"   -> Fill Access Control / {api_id}")

            remove_all_rows_except_first(ws)
            for data_entry in sub_set_data:
                # for key, value in data_entry.items():
                #     print(f"   {key}: {value}")

                row = [
                    data_entry["Required JWT Access Role"],
                    data_entry["Required UAM Data Permission (Access Rights)"]
                ]

                ws.append(row)

            target_api_wb.save(xlsx_file_path)
        else:
            print(f"missing file {xlsx_file_path}")
            pass

    print("------> FILL ACCESS CONTROL DATA END")


def extract_data_from_validation_rules_excel():
    return extract_header_table_data_from_excel(file_validation_rules)


def fill_validation_rules_data_to_each_dist(vr_data_array):
    print("------> FILL VALIDATION RULES DATA START")

    compose_data = compose_module_section_with_api_subset_data(vr_data_array)

    for item in compose_data:
        module_section = item["Module Section"]
        safe_name_module_section = module_section.replace("/", "_")  # handle special character in folder name e.g. "/"
        api_id = item["API ID"]
        sub_set_data = item["Sub Set"]

        module_folder_path = os.path.join(folder_transit_creator, safe_name_module_section)
        xlsx_file_path = os.path.join(module_folder_path, f"{api_id}.xlsx")

        if os.path.exists(xlsx_file_path):
            target_api_wb = openpyxl.load_workbook(xlsx_file_path)

            ws = target_api_wb["VR-TABLE"]

            if not sub_set_data:
                # sub_set_data is empty
                # API No Validation Rules required
                print(f"   -> Fill Validation Rules / {api_id} No Validation Rules")
            else:
                # sub_set_data is not empty
                print(f"   -> Fill Validation Rules / {api_id}")

            remove_all_rows_except_first(ws)
            for data_entry in sub_set_data:
                row = [
                    data_entry["Interface"],
                    data_entry["Entity name"],
                    data_entry["Field name"],
                    data_entry["Field type"],
                    data_entry["Validation Rules"]
                ]

                ws.append(row)

            target_api_wb.save(xlsx_file_path)
        else:
            print(f"missing file {xlsx_file_path}")
            pass

    print("------> FILL VALIDATION RULES DATA END")


def start():
    print(">>>>>>>>>>> EXTRACTOR PROCESS START >>>>>>>>>>>")
    # 1. distribute master data to dist file
    master_data = extract_data_from_master_data_excel()
    fill_master_data_to_each_dist(master_data)

    # 2. extract swagger data and distribute to dist file
    all_file_swag = [lpm.input.Swagger(swag_file) for swag_file in os.listdir(folder_swagger) if
                     swag_file.endswith('.yaml')]

    for current_swg in all_file_swag:
        print("------------------------>    " + os.path.basename(current_swg) + " start")
        current_swagger_data = extract_data_from_swagger(current_swg)
        clear_swagger_data = clear_extract_swagger_data(current_swagger_data)

        fill_swagger_data_to_each_dist(clear_swagger_data)
        print("------------------------>    " + os.path.basename(current_swg) + " end")

    # 3. distribute access control data to dist file
    access_control_data = extract_data_from_access_control_excel()
    fill_access_control_data_to_each_dist(access_control_data)

    # 4. distribute validation rules data to dist file
    validation_rules_data = extract_data_from_validation_rules_excel()
    fill_validation_rules_data_to_each_dist(validation_rules_data)

    print(">>>>>>>>>>> EXTRACTOR PROCESS END >>>>>>>>>>>>>")


if __name__ == '__main__':
    start()
