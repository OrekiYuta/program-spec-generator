import openpyxl
from utils.PathManager import load_path_manager as lpm

file_db_schema = lpm.input("DB-Schema.xlsx")
ping_file = lpm.input("ping.xlsx")
fill_file = lpm.input("fill-ping.xlsx")


def extract_db_schema_data():
    workbook = openpyxl.load_workbook(file_db_schema)

    result_data = []

    for sheet_name in workbook.sheetnames:
        sheet_data = []
        sheet = workbook[sheet_name]

        if sheet.title != "README":

            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data_dict = {}

                all_values_none = True
                for header, value in zip(headers, row):
                    data_dict[header] = value
                    if value is not None:
                        all_values_none = False  # Set the flag to False as long as one value is not None

                # Only set data when all values are not None_ Dict added to sheet_ In data
                if not all_values_none:
                    sheet_data.append(data_dict)

            result_data.append({sheet_name: sheet_data})

    workbook.close()

    # print(result_data)
    return result_data


def clear_db_schema_data(db_data):
    patch_read_data = [
        [
            'Source Table',
            'Destination Field Name',
            'Conversion Logic',
            'DB Column Name(review use only)',
            'Lower Camel Case(review use only)',
            'Actual API Response Name(review use only)'
        ]
    ]

    patch_write_data = [
        [
            'Source Field Name',
            'Destination',
            'Conversion Logic',

            'DB Column Name(review use only)',
            'Lower Camel Case(review use only)',
            'Actual API Response Name(review use only)'
        ]
    ]

    for item in db_data:
        key = list(item.keys())[0]
        value = item[key]

        for column_data in value:
            column_name = column_data['COLUMN_NAME']
            property_name = column_data['PROPERTY']
            api_response_name = column_data['ACTUAL API RESPONSE NAME']

            patch_read_data.append(
                [key.upper(), api_response_name, None, column_name, property_name, api_response_name])
            patch_write_data.append(
                [api_response_name, key.upper(), None, column_name, property_name, api_response_name])

    # for row in patch_read_data:
    #     print(row)

    return patch_read_data, patch_write_data


def fill_data_to_excel(r_data, w_data):
    print("start fill")
    workbook = openpyxl.load_workbook(ping_file)

    sheet = workbook['BRL-DAL-R-Default-Value']

    for row_data in r_data:
        sheet.append(row_data)

    sheet = workbook['BRL-DAL-W-Default-Value']
    for row_data in w_data:
        sheet.append(row_data)

    workbook.save(fill_file)
    print("finish fill")


if __name__ == '__main__':
    db_schema_data = extract_db_schema_data()
    p_read_data, p_write_data = clear_db_schema_data(db_schema_data)
    fill_data_to_excel(p_read_data, p_write_data)
