import os
import openpyxl
import yaml

from extractor import extract_data_from_swagger, clear_extract_swagger_data
from utils.PathManager import load_path_manager as lpm

file_master_data = lpm.input("ping.xlsx") 
file_swg_common_aa = lpm.input("common-swagger-aa.yaml")
file_swg_common_sys = lpm.input("common-swagger-system.yaml")
file_swg_batch_sys = lpm.input("batch-swagger-system.yaml")
file_swg_cc_sys = lpm.input("cc-swagger-system.yaml")


def ping_count(swg_data):
    wb = openpyxl.load_workbook(file_master_data)
    sheet = wb["MASTER"]

    excel_headers = [cell.value for cell in sheet[1]]

    api_endpoint_column_index = excel_headers.index("API End Point")
    http_method_column_index = excel_headers.index("HTTP Method")
    module_section_column_index = excel_headers.index("Module Section")
    api_id_column_index = excel_headers.index("API ID")

    all_master_excel_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # if row[api_id_column_index].startswith("AA-API"):
        all_master_excel_rows.append(row[api_id_column_index])

    ping_master_excel_rows = []
    ping_swg_rows = []  # ping if all api

    for swg_item_data in swg_data:
        api_endpoint = swg_item_data["API End Point"]
        http_method = swg_item_data["HTTP Method"]

        matching_rows = []

        current_module_section = None

        for row in sheet.iter_rows(min_row=2, values_only=True):

            #  handle module section
            if row[module_section_column_index] is not None:
                current_module_section = row[module_section_column_index]

            # print("search:  " + http_method + "   " + api_endpoint.replace("/internal", ""))
            # print("target:  " + row[http_method_column_index] + "   " + row[api_endpoint_column_index])

            if row[http_method_column_index].lower().strip() == http_method.lower().strip():
                if (
                        api_endpoint.lower().strip().replace("/internal", "") ==
                        row[api_endpoint_column_index].lower().strip().replace(" ", "")
                                .replace("/cas/aa/v1", "")  # common swg
                                .replace("/cas/aa/batch/v1/internal", "")  # batch swg
                                .replace("/cas/cc/v1", "")  # cvh swg
                                # .replace("/v1", "")
                ):
                    # print("=> ping:    " + row[api_id_column_index] + "   "
                    #       + row[http_method_column_index] + "   "
                    #       + row[api_endpoint_column_index])

                    row_list = list(row)
                    ping_master_excel_rows.append(row[api_id_column_index])

                    row_list[module_section_column_index] = current_module_section  # handle module section
                    matching_rows.append(tuple(row_list))

        if matching_rows:
            ping_swg_rows.append(matching_rows[0][api_id_column_index])  # ping if all api
        else:
            print("=> un-ping:    " + http_method + "     " + api_endpoint)

    # print("=> " + module_section + " => " + api_id)
    print("ping_swg_rows    " + str(len(ping_swg_rows)))
    # print(ping_swg_rows)

    print("all_master_excel_rows    " + str(len(all_master_excel_rows)) + " " + str(all_master_excel_rows))
    print("ping_master_excel_rows   " + str(len(ping_master_excel_rows)) + " " + str(ping_master_excel_rows))

    all_master_set = set(all_master_excel_rows)
    ping_master_set = set(ping_master_excel_rows)

    un_ping_master_set = all_master_set - ping_master_set

    print("un_ping_master_excel_rows   " + str(len(un_ping_master_set)) + " " + str(list(un_ping_master_set)))
    wb.close()


def check_api_mapping_method():
    all_file_swag = [
        file_swg_common_aa,
        file_swg_common_sys,
        file_swg_batch_sys,
        file_swg_cc_sys
    ]

    for current_swg in all_file_swag:
        print("------------------------>    " + os.path.basename(current_swg))
        current_swagger_data = extract_data_from_swagger(current_swg)
        clear_swagger_data = clear_extract_swagger_data(current_swagger_data)
        ping_count(clear_swagger_data)


def check_api_url_count():
    all_file_swag = [
        file_swg_common_aa,
        file_swg_common_sys,
        file_swg_batch_sys,
        file_swg_cc_sys
    ]

    for current_swg in all_file_swag:
        print("------------------------>    " + os.path.basename(current_swg))
        current_swagger_data = extract_data_from_swagger(current_swg)

        count = 0
        for url, data in current_swagger_data.items():
            count += len(data)
            print(url)

        print(os.path.basename(current_swg) + "     " + str(count))


check_api_mapping_method()
