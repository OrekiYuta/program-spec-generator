import openpyxl
from docx import Document
from openpyxl.styles import Font

from utils.PathManager import load_path_manager as lpm

file_master_data = lpm.input("ping.xlsx")
output_file = lpm.input("API Test Master Data.xlsx")
api_inve_aa = lpm.refer("API Inventory\AA.xlsx")
api_inve_bb = lpm.refer("API Inventory\BB.xlsx")
api_inve_cc = lpm.refer("API Inventory\CC.xlsx")
file_access_control = lpm.input("Access Control.xlsx")


def extract_data_from_master_data_excel():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["MASTER"]

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    data_list = []

    current_module_section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_map = {}
        for col_idx, value in enumerate(row, start=1):

            col_name = header[col_idx - 1]

            # assort module section
            if col_idx == 1:
                if value is not None:
                    current_module_section = value
                else:
                    value = current_module_section

            data_map[col_name] = value

        data_list.append(data_map)

    wb.close()
    return data_list


def generate_api_test_master_data_excel(clear_data, master_data):
    wb = openpyxl.Workbook()

    print("fill in excel")
    count = 0
    exist_api_sheet = []
    for item in clear_data:
        api_id = item.get("API ID", "Unknown")
        print(f"fill {api_id}")

        if api_id not in exist_api_sheet:
            sheet = wb.create_sheet(api_id)
            exist_api_sheet.append(api_id)

            sheet.append([
                "FS ID",
                "API Index",
                "API ID",
                "HTTP Method",
                "API End Point",
                "Function ID",
                "Function Name",
                "Description",
                "Role Required",
                "Scenario",
                "Test Description",
                "Test Case Logic",
                "Corresponding Fields",
                "Request API URL",
                "Request Header",
                "Path Parameter(s)",
                "Request Body",
                "Tester",
                "Testing Date",
                "Reviewer",
                "Review Date",
                # ------------------
                "Contents",
                "Response Body (RAW)",
                "Response Status",
                "Response File Path",
                "Response File",
                "Trace ID",
                "SQL Path",
                "SQL",
                "Jaeger Screenshot Path",
                "Jaeger Screenshot",
                "JUnit Test Screenshot Path",
                "JUnit Test Screenshot",
                "DB Schema Screenshot Path",
                "DB Schema Screenshot"
            ])

            bold_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = bold_font

            data_array = [item.get(key, '') if item.get(key) is not None else '' for key in item.keys()]
            sheet.append(data_array)

        else:
            sheet = wb[api_id]
            data_array = [item.get(key, '') if item.get(key) is not None else '' for key in item.keys()]
            sheet.append(data_array)
        count += 1

    readme_ws = wb.create_sheet("README", 0)  # Add README sheet at the beginning
    readme_ws.append(["Module Section", "API Section", "API Index", "API ID", "action"])

    for ic in master_data:
        module_section = ic.get("Module Section", "Unknown")
        api_section = ic.get("API Section", "Unknown")
        api_index = ic.get("API Index", "Unknown")
        api_id = ic.get("API ID", "Unknown")
        action = f'=HYPERLINK("#\'" & D2 & "\'!D1", "Go to Sheet " & D2)'
        readme_ws.append([module_section, api_section, api_index, api_id, action])

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_file)
    print(f"finish {count}")


def extract_api_inventory_data(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb['RESTful API']

    header_row = sheet[1]
    headers = {header.value: index for index, header in enumerate(header_row, start=0)}

    columns_to_extract = ["Internal Function ID", "Function", "API ID"]

    result_data = []
    current_api_id = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_item = {column: row[headers[column]] for column in columns_to_extract}
        if data_item["API ID"] is None:
            data_item["API ID"] = current_api_id
        else:
            current_api_id = data_item["API ID"]
        result_data.append(data_item)

    wb.close()
    return result_data


def clear_api_inventory_data():
    iv_aa = extract_api_inventory_data(api_inve_aa)
    iv_bb = extract_api_inventory_data(api_inve_bb)
    iv_cc = extract_api_inventory_data(api_inve_cc)

    compose_data = iv_aa + iv_bb + iv_cc

    api_test_data = []
    api_test_data_item = {
        "FS ID": "TODO",
        "API Index": "TODO",
        "API ID": "TODO",
        "HTTP Method": "TODO",
        "API End Point": "TODO",
        "Function ID": "TODO",
        "Function Name": "TODO",
        "Description": "TODO",
        "Role Required": "TODO",
        "Scenario": "TODO",
        "Test Description": "TODO",
        "Test Case Logic": "TODO",
        "Corresponding Fields": "TODO",
        "Request API URL": "TODO",
        "Request Header": None,
        "Path Parameter(s)": "TODO",
        "Request Body": "TODO",
        "Tester": "TODO",
        "Testing Date": "TODO",
        "Reviewer": "TODO",
        "Review Date": "TODO",
        # ------------------
        "Contents": None,
        "Response Body (RAW)": None,
        "Response Status": None,
        "Response File Path": None,
        "Response File": None,
        "Trace ID": None,
        "SQL Path": None,
        "SQL": None,
        "Jaeger Screenshot Path": None,
        "Jaeger Screenshot": None,
        "JUnit Test Screenshot Path": None,
        "JUnit Test Screenshot": None,
        "DB Schema Screenshot Path": None,
        "DB Schema Screenshot": None
    }

    for ms in compose_data:
        current_item = api_test_data_item.copy()
        api_id = ms.get("API ID", "TODO")
        print(f"process {api_id}")
        if api_id == "-":
            continue

        item = get_master_data_item(api_id)
        role = get_role_from_access_control_excel(api_id)
        current_item["FS ID"] = item.get("Module Section", "TODO")
        current_item["API Index"] = item.get("API Index", "TODO")
        current_item["HTTP Method"] = item.get("HTTP Method", "TODO")
        current_item["API End Point"] = item.get("API End Point", "TODO")
        current_item["Description"] = item.get("Description", "TODO")

        current_item["API ID"] = ms.get("API ID", "TODO")
        current_item["Function ID"] = ms.get("Internal Function ID", "TODO")
        current_item["Function Name"] = ms.get("Function", "TODO")
        current_item["Role Required"] = role
        current_item["Request API URL"] = "https://" + item.get("API End Point", "TODO")

        api_test_data.append(current_item)

    sorted_api_test_data = sorted(api_test_data, key=lambda x: x['API Index'])

    return sorted_api_test_data


def get_master_data_item(api_id):
    master_data = extract_data_from_master_data_excel()

    target_item = None
    for ms in master_data:
        if ms["API ID"] == api_id:
            target_item = ms
            break

    return target_item


def get_role_from_access_control_excel(api_id):
    workbook = openpyxl.load_workbook(file_access_control)

    result_data = []

    for sheet_name in workbook.sheetnames:
        sheet_data = []
        sheet = workbook[sheet_name]

        if sheet.title != "README":

            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data_dict = {}
                for header, value in zip(headers, row):
                    data_dict[header] = value
                sheet_data.append(data_dict)

            result_data.append({sheet_name: sheet_data})

    workbook.close()

    required_jwt_access_role = None
    matching_item = next((item for item in result_data if api_id in item), None)

    if matching_item:
        required_jwt_access_role = matching_item[api_id][0]['Required JWT Access Role']
        print(required_jwt_access_role)
    else:
        print("API ID not found in the data")

    return required_jwt_access_role


if __name__ == '__main__':
    ms_data = extract_data_from_master_data_excel()
    cl_data = clear_api_inventory_data()
    generate_api_test_master_data_excel(cl_data, ms_data)
