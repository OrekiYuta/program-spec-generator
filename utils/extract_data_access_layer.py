import docx
import openpyxl
from docx import Document
from utils.PathManager import load_path_manager as lpm

MASTER_DATA_FILE = "ping.xlsx"
file_refer = lpm.refer("Program Specification.docx")
input_file = lpm.input(MASTER_DATA_FILE)
output_file = lpm.input("Filled-" + MASTER_DATA_FILE)


def fill_data_to_master_data_excel(data):
    input_file = lpm.input(MASTER_DATA_FILE)
    output_file = lpm.input("Filled-" + MASTER_DATA_FILE)

    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook['MASTER']
    print("Fill to excel")
    # Create a mapping of column names to their indices
    col_name_to_idx = {}
    for col_idx, col_name in enumerate(sheet[1], start=1):
        col_name_to_idx[col_name.value] = col_idx

    count = 0
    for api_item in data:
        # Find the matching row
        api_section = api_item['API Section']
        api_section_col_idx = col_name_to_idx.get('API Section')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[api_section_col_idx - 1] == api_section:
                print(f"fill {api_section}")
                count += 1
                # Update data in the matching row
                for key, value in api_item.items():
                    if key != 'API Section':
                        col_idx = col_name_to_idx.get(key)
                        if col_idx:
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                break  # Exit loop after finding the match

    print(f"finish {count}")
    workbook.save(output_file)


def extract_api_read_write_mark():
    doc = Document(file_refer)

    api_sections = []
    current_section = None
    api_section_data = {}
    table_count = len(doc.tables)  # Note
    # print(table_count)
    for i, element in enumerate(doc.element.body):
        if isinstance(element, docx.oxml.text.paragraph.CT_P):
            paragraph = docx.text.paragraph.Paragraph(element, doc)
            text = paragraph.text.strip()

            if text.startswith("AA-") or text.startswith("BB-"):
                if current_section:
                    api_sections.append(api_section_data)
                    api_section_data = {}

                current_section = text
                api_section_data["API Section"] = current_section
                # print(current_section)

            elif text.startswith("API ID"):
                api_section_data["API ID"] = text.split(":")[-1].strip()
                print("----------------------------------------------------------")
                print(text.split(":")[-1].strip())
            elif text.startswith("Operation:"):
                print(text)
                if "Data Access Layer Operation" not in api_section_data:
                    api_section_data["Data Access Layer Operation"] = text.replace("Operation:", "").strip()
                else:
                    api_section_data["Data Access Layer Operation"] += "," + text.replace("Operation:", "").strip()

        elif isinstance(element, docx.oxml.table.CT_Tbl):
            table = docx.table.Table(element, doc)
            # print("ping table")
            # print(len(table.rows))

            first_row = table.rows[0]
            if len(first_row.cells) > 1:
                first_cell = first_row.cells[0]
                second_cell = first_row.cells[1]
                first_cell_text = first_cell.text.strip()
                second_cell_text = second_cell.text.strip()

                if first_cell_text == "Source Table" and second_cell_text == "Source Field Name":
                    # print("read table")
                    source_table_data = []
                    for row in table.rows[1:]:
                        value = row.cells[0].text.strip()
                        if value and value.upper() != "N/A":
                            source_table_data.append(value)

                    unique_source_table_data = list(set(source_table_data))
                    unique_upper_source_table_data = [value.upper() for value in unique_source_table_data]
                    unique_read_table_values = ", ".join(unique_upper_source_table_data)
                    api_section_data["Database Operation Table - Read"] = unique_read_table_values

                    print(f"   <Read> Table Table Name: {unique_upper_source_table_data}")

                elif first_cell_text == "Source" and second_cell_text == "Source Field Name":
                    # print("write table")
                    destination_data = []
                    for row in table.rows[1:]:
                        value = row.cells[2].text.strip()
                        if value and value.upper() != "N/A":
                            destination_data.append(value)

                    unique_destination_data = list(set(destination_data))
                    unique_upper_destination_data = [value.upper() for value in unique_destination_data]
                    unique_write_table_values = ", ".join(unique_upper_destination_data)
                    api_section_data["Database Operation Table - Write"] = unique_write_table_values

                    print(f"   <Write> Table Table Name: {unique_upper_destination_data}")
        else:
            continue

    if current_section:
        api_sections.append(api_section_data)

    # print(api_sections)
    return api_sections


if __name__ == '__main__':
    data_access_layer_data = extract_api_read_write_mark()
    fill_data_to_master_data_excel(data_access_layer_data)
