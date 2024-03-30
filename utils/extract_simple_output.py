import docx
import openpyxl
from docx import Document
from utils.PathManager import load_path_manager as lpm

MASTER_DATA_FILE = "ping.xlsx"
file_refer = lpm.refer("Program Specification.docx")
input_file = lpm.input(MASTER_DATA_FILE)
output_file = lpm.input("Filled-" + MASTER_DATA_FILE)


def fill_data_to_master_data_excel(data):
    input_file = lpm.input(MASTER_DATA_FILE)
    output_file = lpm.input("Filled-" + MASTER_DATA_FILE)

    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook['MASTER']
    print("Fill to excel")
    # Create a mapping of column names to their indices
    col_name_to_idx = {}
    for col_idx, col_name in enumerate(sheet[1], start=1):
        col_name_to_idx[col_name.value] = col_idx

    count = 0
    for api_item in data:
        # Find the matching row
        api_section = api_item['API Section']
        api_section_col_idx = col_name_to_idx.get('API Section')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[api_section_col_idx - 1] == api_section:
                print(f"fill {api_section}")
                count += 1
                # Update data in the matching row
                for key, value in api_item.items():
                    if key != 'API Section':
                        col_idx = col_name_to_idx.get(key)
                        if col_idx:
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                break  # Exit loop after finding the match

    print(f"finish {count}")
    workbook.save(output_file)


def extract_api_sample_output():
    doc = Document(file_refer)

    api_sections = []
    api_section_data = {}
    current_section = None
    table_count = len(doc.tables)  # Note
    # print(table_count)
    count = 0
    sample_flag = False

    all_api = []
    current_api = None
    ping_sp_api = []

    for i, element in enumerate(doc.element.body):
        # print(f"test {sample_flag}")
        if isinstance(element, docx.oxml.text.paragraph.CT_P):
            paragraph = docx.text.paragraph.Paragraph(element, doc)
            text = paragraph.text.strip()

            if text.startswith("AA-") or text.startswith("BB-"):
                sample_flag = False
                if current_section:
                    api_sections.append(api_section_data)
                    api_section_data = {}

                current_section = text
                api_section_data["API Section"] = current_section
                print("----------------------------------------------------------")
                print(current_section)

            elif text.startswith("API ID"):

                api_section_data["API ID"] = text.split(":")[-1].strip()

                print(text.split(":")[-1].strip())
                current_api = text.split(":")[-1].strip()
                all_api.append(text.split(":")[-1].strip())

            elif text.startswith("Sample Output"):
                print(text)
                print("ping sample output")
                sample_flag = True

        elif isinstance(element, docx.oxml.table.CT_Tbl):
            if sample_flag is True:
                table = docx.table.Table(element, doc)
                print(len(table.rows))
                first_row = table.rows[0]
                # print(first_row.cells[0])

                if len(first_row.cells) == 1:
                    print("ping sample output table")
                    count += 1
                    first_cell = first_row.cells[0]

                    api_section_data["Sample Output"] = first_cell.text

                    # print(f"   Sample Output : {first_cell}")
                    sample_flag = False
                    ping_sp_api.append(current_api)
                else:
                    print("after flag , is sample output? but no 1 row")
                    pass
            else:
                # print("is table, but not is sample output table")
                pass
        else:
            if sample_flag is True:
                print("is sample output format , but not is a word format table , only look like table")
            continue

    if current_section:
        api_sections.append(api_section_data)

    # print(api_sections)
    print("-------------------------")
    print(f"finish {count}")
    un_ping_sp_api = [api for api in all_api if api not in ping_sp_api]
    print(f"un ping {un_ping_sp_api}")
    return api_sections


if __name__ == '__main__':
    data = extract_api_sample_output()
    fill_data_to_master_data_excel(data)
