import os
import xml.etree.ElementTree as ET
from collections import Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from utils.PathManager import load_path_manager as lpm


def extract_single_xml_schema_data(file):
    tree = ET.parse(file)
    root = tree.getroot()

    data_array = []

    '''
        get all <resultMap/> e.g.
        <resultMap id="BaseResultMap" />
        <resultMap extends="BaseResultMap" />
    '''
    # for result_map in root.findall(".//resultMap"):

    # only get <resultMap id="BaseResultMap" />
    for result_map in root.findall(".//resultMap[@id='BaseResultMap']"):
        #  <id column="ID" jdbcType="VARBINARY" property="id" />
        id_element = result_map.find(".//id")
        if id_element is not None:
            id_column = id_element.get('column')
            id_jdbc_type = id_element.get('jdbcType')
            property_ = id_element.get('property')
            data_array.append({
                'column': id_column,
                'jdbcType': id_jdbc_type,
                'property': property_,
            })

        # <result column="ANNOUNCE_SYS" jdbcType="VARCHAR" property="announceSys" />
        for result in result_map.findall(".//result"):
            id_column = result.get('column')
            id_jdbc_type = result.get('jdbcType')
            property_ = result.get('property')
            data_array.append({
                'column': id_column,
                'jdbcType': id_jdbc_type,
                'property': property_,
            })

    return data_array


def extract_single_xml_table_name(file):
    tree = ET.parse(file)
    root = tree.getroot()

    select_element = root.find(".//select[@id='selectByPrimaryKey']")

    if select_element is not None:
        select_sql = ET.tostring(select_element, encoding='unicode')

        from_index = select_sql.lower().find("from")
        where_index = select_sql.lower().find("where")

        if from_index != -1 and where_index != -1:
            from_clause = select_sql[from_index + 4:where_index].strip()
            table_name = from_clause.split()[0]
            print("Table name:" + table_name + "    " + os.path.basename(file))
            return table_name
        else:
            print("Could not extract table name from the SQL statement")
            return None
    else:
        print("Could not extract selectByPrimaryKey tag     " + os.path.basename(file))
        return None


# Deprecated
def fill_data_to_excel():
    file_xml = lpm.refer.Mybatis.bb_cas_cvh("PortalConvsnMapper.xml")
    table_name = extract_single_xml_table_name(file_xml)
    if table_name is not None:
        xml_data = extract_single_xml_schema_data(file_xml)

        wb = Workbook()
        ws = wb.active
        ws.title = table_name

        ws.append(['COLUMN_NAME', 'DATA_TYPE'])

        for item in xml_data:
            ws.append([item['column'], item['jdbcType']])

        wb.save('output.xlsx')


def process_generate_excel(folder_path):
    file_excel_name = os.path.basename(folder_path)
    sheet_count = 0
    wb = Workbook()
    for filename in os.listdir(folder_path):
        if filename.endswith(".xml"):
            file_xml = os.path.join(folder_path, filename)

            table_name = extract_single_xml_table_name(file_xml)

            if table_name is not None:

                xml_data = extract_single_xml_schema_data(file_xml)

                ws = wb.create_sheet(title=table_name)
                sheet_count += 1

                ws.append(
                    [
                        'COLUMN_NAME',
                        'DATA_TYPE',
                        'PROPERTY'
                    ]
                )
                # bold header/first row
                bold_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = bold_font

                for item in xml_data:
                    ws.append(
                        [
                            item['column'],
                            item['jdbcType'],
                            item['property'],
                        ]
                    )

            else:
                # to next xml file
                continue

    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    file_excel = f"db-schema-{file_excel_name}.xlsx"
    wb.save(lpm.input(file_excel))
    print("sheet count  " + str(sheet_count))


def gen_compose_dist_file():
    sheet_count = 0

    folder_bb_cas_cvh = str(lpm.refer.Mybatis.bb_cas_cvh)
    folder_bb_cas_aa_batch = str(lpm.refer.Mybatis.bb_cas_aa_batch)
    folder_bb_cas_aa_common = str(lpm.refer.Mybatis.bb_cas_aa_common)

    file_array = [
        folder_bb_cas_aa_common,
        folder_bb_cas_aa_batch,
        folder_bb_cas_cvh
    ]

    wb = Workbook()
    original_list = []

    for current_file in file_array:

        for filename in os.listdir(current_file):
            if filename.endswith(".xml"):
                file_xml = os.path.join(current_file, filename)

                table_name = extract_single_xml_table_name(file_xml)

                if table_name is not None:
                    if table_name not in original_list:
                        xml_data = extract_single_xml_schema_data(file_xml)

                        ws = wb.create_sheet(title=table_name)
                        sheet_count += 1

                        ws.append(
                            [
                                'COLUMN_NAME',
                                'DATA_TYPE',
                                'PROPERTY',
                                'ACTUAL API RESPONSE NAME'
                            ]
                        )
                        # bold header/first row
                        bold_font = Font(bold=True)
                        for cell in ws[1]:
                            cell.font = bold_font

                        for item in xml_data:
                            ws.append(
                                [
                                    item['column'],
                                    item['jdbcType'],
                                    item['property'],
                                    item['property']  # ACTUAL API RESPONSE NAME , Then manual check it
                                ]
                            )

                        print("add " + table_name)
                    else:
                        print("exist and skip " + table_name)
                else:
                    # to next xml file
                    continue

                original_list.append(table_name)

    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    file_excel = f"DB-Schema.xlsx"
    wb.save(lpm.input(file_excel))
    print("sheet count  " + str(sheet_count))
    counter = Counter(original_list)
    unique_list = list(counter.keys())
    print(unique_list)
    print(len(unique_list))


def gen_each_project_dist_file():
    folder_bb_cas_cvh = str(lpm.refer.Mybatis.bb_cas_cvh)
    folder_bb_cas_aa_batch = str(lpm.refer.Mybatis.bb_cas_aa_batch)
    folder_bb_cas_aa_common = str(lpm.refer.Mybatis.bb_cas_aa_common)

    process_generate_excel(folder_bb_cas_cvh)
    process_generate_excel(folder_bb_cas_aa_batch)
    process_generate_excel(folder_bb_cas_aa_common)


def get_table_name(folder_path):
    table_name_list = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".xml"):
            file_xml = os.path.join(folder_path, filename)

            table_name = extract_single_xml_table_name(file_xml)

            if table_name is not None:
                table_name_list.append(table_name)

            else:
                # to next xml file
                continue

    # print(table_name_list)
    return table_name_list


def get_all_table_name():
    folder_bb_cas_cvh = str(lpm.refer.Mybatis.bb_cas_cvh)
    folder_bb_cas_aa_batch = str(lpm.refer.Mybatis.bb_cas_aa_batch)
    folder_bb_cas_aa_common = str(lpm.refer.Mybatis.bb_cas_aa_common)
    cvh = get_table_name(folder_bb_cas_cvh)
    batch = get_table_name(folder_bb_cas_aa_batch)
    common = get_table_name(folder_bb_cas_aa_common)

    merged_set = set(cvh + batch + common)
    merged_list = list(merged_set)
    print(merged_list)
    print(len(merged_list))


if __name__ == '__main__':
    gen_compose_dist_file()
    # get_all_table_name()
