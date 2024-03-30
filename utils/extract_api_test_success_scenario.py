import openpyxl
from utils.PathManager import load_path_manager as lpm

file_api_test = lpm.refer("API Test Master Data.xlsx")
file_master_data = lpm.input("ping.xlsx")


def extract_success_scenario_data_from_api_test_excel():
    workbook = openpyxl.load_workbook(file_api_test)

    result_data = []

    for sheet_name in workbook.sheetnames:
        sheet_data = []
        sheet = workbook[sheet_name]

        if sheet.title != "README" and sheet.title != "REFER":
            ping = False
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):

                data_dict = {}
                for header, value in zip(headers, row):
                    data_dict[header] = value

                if "Positive Scenario 1" in data_dict['Scenario']:
                    ping = True
                    sheet_data.append(data_dict)
                    break
                elif "200 OK" in data_dict['Test Case Logic']:
                    ping = True
                    sheet_data.append(data_dict)
                    break
                else:
                    # print(f"{sheet_name} -  {data_dict['Scenario']}")
                    pass
            if ping is False:
                # print(sheet_name)
                # print(f"{sheet_name} -  {data_dict['Scenario']}")
                # print("-----------------------")
                pass

            result_data.append({sheet_name: sheet_data})

    workbook.close()

    return result_data


def load_master_data_sample_output_sheet():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["Sample Output"]
    table_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            table_data.append(row)
    wb.close()
    return table_data


def compose_sample_output_data(ms_data, sn_data):
    for index, item in enumerate(ms_data):
        api_id, http_method, api_endpoint, _, _, run_it, last_run_date, result, trace_id, response_body = item

        for sn_item in sn_data:
            if sn_item.get(api_id):
                sn_item_values = sn_item[api_id]
                for sn_item_value in sn_item_values:
                    path_parameters = sn_item_value.get('Path Parameter(s)')
                    request_body = sn_item_value.get('Request Body')

                    master_data[index] = (api_id, http_method, api_endpoint, path_parameters,
                                          request_body, run_it, last_run_date, result, trace_id, response_body)
                    break

    return ms_data


def fill_master_data_sample_output_pp_rb_col(comp_data):
    wb = openpyxl.load_workbook(file_master_data)
    sheet = wb["Sample Output"]
    for item in comp_data:
        ms_id, _, _, _, _, _, _, _, _, _ = item

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == ms_id:
                    sheet.cell(row=cell.row, column=4, value=item[3])
                    sheet.cell(row=cell.row, column=5, value=item[4])
                    sheet.cell(row=cell.row, column=6, value=item[5])
                    sheet.cell(row=cell.row, column=7, value=item[6])
                    sheet.cell(row=cell.row, column=8, value=item[7])
                    sheet.cell(row=cell.row, column=9, value=item[8])
                    sheet.cell(row=cell.row, column=10, value=item[9])

    wb.save(file_master_data)
    print("finish")


if __name__ == '__main__':
    master_data = load_master_data_sample_output_sheet()
    scenario_data = extract_success_scenario_data_from_api_test_excel()
    compose_data = compose_sample_output_data(master_data, scenario_data)
    fill_master_data_sample_output_pp_rb_col(compose_data)
