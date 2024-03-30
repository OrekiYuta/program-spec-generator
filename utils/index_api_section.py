import openpyxl
from docx import Document
from utils.PathManager import load_path_manager as lpm

file_refer = lpm.refer("Program Specification.docx")
file_master_data = lpm.input("ping.xlsx")
file_fill_master_data = lpm.input("fill-ping.xlsx")


def extract_data_from_prog_spec_word_index():
    doc = Document(file_refer)

    api_sections = []
    current_section = None
    api_section_data = {}
    count = 0
    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()

        if text.startswith("AA-") or text.startswith("BB-"):
            if current_section:
                api_sections.append(api_section_data)
                api_section_data = {}

            current_section = text
            api_section_data["API Section"] = current_section

        elif text.startswith("API ID"):
            api_section_data["API ID"] = paragraph.text.split(":")[-1].strip()
            count += 1
            api_section_data["API Index"] = count

    if current_section:
        api_sections.append(api_section_data)

    print(api_sections)

    return api_sections


def extract_data_from_master_data_excel():
    wb = openpyxl.load_workbook(file_master_data)
    ws = wb["MASTER"]

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    data_list = []

    current_module_section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        data_map = {}
        for col_idx, value in enumerate(row, start=1):

            col_name = header[col_idx - 1]

            # assort module section
            if col_idx == 1:
                if value is not None:
                    current_module_section = value
                else:
                    value = current_module_section

            data_map[col_name] = value

        data_list.append(data_map)

    wb.close()
    # print(data_list)
    return data_list


def append_index_columns(word_index, master_data_index):
    for master_item in master_data_index:
        master_api_id = master_item['API ID']
        for word_item in word_index:
            word_api_id = word_item['API ID']
            if word_api_id == master_api_id:
                master_item['API Index'] = word_item['API Index']
                break

    return master_data_index


def fill_data_to_master_data_excel(api_data):
    input_file = lpm.input(file_master_data)
    output_file = lpm.input(file_fill_master_data)

    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook['MASTER']
    print("Fill to excel")

    # Create a mapping of column names to their indices
    col_name_to_idx = {}
    for col_idx, col_name in enumerate(sheet[1], start=1):
        col_name_to_idx[col_name.value] = col_idx

    count = 0
    for api_item in api_data:
        api_section = api_item['API Section']
        api_section_col_idx = col_name_to_idx.get('API Section')

        # Find the matching row
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[api_section_col_idx - 1] == api_section:
                print(f"fill {api_section}")
                count += 1
                # Update data in the matching row
                for key, value in api_item.items():
                    if key != 'API Section':
                        col_idx = col_name_to_idx.get(key)
                        if col_idx:
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                break  # Exit loop after finding the match

    print(f"finish {count}")
    workbook.save(output_file)


if __name__ == '__main__':
    word_data = extract_data_from_prog_spec_word_index()
    master_data = extract_data_from_master_data_excel()
    result_index = append_index_columns(word_data, master_data)
    fill_data_to_master_data_excel(result_index)
