import openpyxl
from docx import Document

from utils.PathManager import load_path_manager as lpm

file_default = lpm.input("default.xlsx")
file_ping = lpm.input("ping.xlsx")
fill_file = lpm.input("fill-ping.xlsx")


def extract_default_value(file_path):
    workbook = openpyxl.load_workbook(file_path)

    result_data = []

    for sheet_name in workbook.sheetnames:
        sheet_data = []
        sheet = workbook[sheet_name]

        if sheet.title == "BRL-DAL-R-Default-Value" or sheet.title == "BRL-DAL-W-Default-Value":

            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data_dict = {}

                all_values_none = True
                for header, value in zip(headers, row):
                    data_dict[header] = value
                    if value is not None:
                        all_values_none = False  # Set the flag to False as long as one value is not None

                # Only set data when all values are not None_ Dict added to sheet_ In data
                if not all_values_none:
                    sheet_data.append(data_dict)

            result_data.append({sheet_name: sheet_data})

    workbook.close()

    return result_data


def extract_master_data_default_value():
    return extract_default_value(file_ping)


def extract_default_value_pre():
    return extract_default_value(file_default)


def compose_default_value_data():
    ...


if __name__ == '__main__':
    ms_df_data = extract_master_data_default_value()
    pre_df_data = extract_default_value_pre()
    compose_default_value_data()
