import openpyxl
from docx import Document
from utils.PathManager import load_path_manager as lpm

MASTER_DATA_FILE = "Program Spec Master Data.xlsx"
file_refer = lpm.refer("Program Specification.docx")


def extract_data_from_prog_spec_word():
    doc = Document(file_refer)

    api_sections = []
    current_section = None
    api_section_data = {}

    get_next_line_references = False
    get_next_line_header_parameters = False
    get_next_line_business_logic = False

    get_next_line_service = False
    get_next_line_controller = False

    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()

        if text.startswith("AA-") or text.startswith("BB-"):
            if current_section:
                api_sections.append(api_section_data)
                api_section_data = {}

            current_section = text
            api_section_data["API Section"] = current_section

            get_next_line_references = False
            get_next_line_header_parameters = False
            get_next_line_business_logic = False

        elif text.startswith("Description"):
            api_section_data["Description"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("Type"):
            api_section_data["Type"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("Service"):
            if text.split(":")[0].strip() == "Service":
                api_section_data["Service"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("Mode"):
            api_section_data["Mode"] = paragraph.text.split(":")[-1].strip()

        elif text.startswith("Controller"):

            # handle error typo
            value = paragraph.text.split(":")[-1].strip()
            if not value.startswith("/"):
                value = "/" + value.replace("\\", "/")

            api_section_data["Controller"] = value
            get_next_line_controller = True
        elif get_next_line_controller:
            if text.startswith("Servicx") is False:
                text = text.replace("\\", "/")  # handle error typo

                api_section_data["Controller"] += text.strip()
                get_next_line_controller = False

        # Note Program Spec word rename the Service(s) to Servicx
        elif text.startswith("Servicx"):
            # if text.split(":")[0].strip() == "Service(s)":
            api_section_data["Service(s)"] = paragraph.text.split(":")[-1].strip()
            get_next_line_service = True
        elif get_next_line_service:
            if text.startswith("API End Point") is False:
                api_section_data["Service(s)"] += text.strip()
                get_next_line_service = False

        elif text.startswith("API End Point"):
            api_section_data["API End Point"] = paragraph.text.split(":")[-1].strip()

        elif text.startswith("HTTP Method"):
            api_section_data["HTTP Method"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("API Service Method"):
            api_section_data["API Service Method"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("API ID"):
            api_section_data["API ID"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("Consumer"):
            api_section_data["Consumer"] = paragraph.text.split(":")[-1].strip()
        elif text.startswith("API Specification Version"):
            api_section_data["API Specification Version"] = paragraph.text.split(":")[-1].strip()

        elif text == "References":
            get_next_line_references = True
        elif get_next_line_references:
            api_section_data["References"] = "N/A" if text is None else text
            get_next_line_references = False
        elif text == "Header Parameter(s)":
            get_next_line_header_parameters = True
        elif get_next_line_header_parameters:
            api_section_data["Header Parameter(s)"] = text
            get_next_line_header_parameters = False
        elif text == "Business Rules & Logic (Description / Pseudo Code)":
            get_next_line_business_logic = True
        elif get_next_line_business_logic:
            api_section_data["Business Rules & Logic - Abstract"] = text
            get_next_line_business_logic = False

    if current_section:
        api_sections.append(api_section_data)

    print(api_sections)

    return api_sections


def fill_data_to_master_data_excel(api_data):
    input_file = lpm.input(MASTER_DATA_FILE)
    output_file = lpm.input("Filled-" + MASTER_DATA_FILE)

    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook['MASTER']

    # Create a mapping of column names to their indices
    col_name_to_idx = {}
    for col_idx, col_name in enumerate(sheet[1], start=1):
        col_name_to_idx[col_name.value] = col_idx

    for api_item in api_data:
        api_section = api_item['API Section']
        api_section_col_idx = col_name_to_idx.get('API Section')

        # Find the matching row
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[api_section_col_idx - 1] == api_section:
                # Update data in the matching row
                for key, value in api_item.items():
                    if key != 'API Section':
                        col_idx = col_name_to_idx.get(key)
                        if col_idx:
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                break  # Exit loop after finding the match

    workbook.save(output_file)


if __name__ == '__main__':
    master_data = extract_data_from_prog_spec_word()
    fill_data_to_master_data_excel(master_data)
