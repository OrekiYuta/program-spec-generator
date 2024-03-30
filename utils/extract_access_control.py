import docx
import openpyxl
from docx import Document
from openpyxl.styles import Font

from utils.PathManager import load_path_manager as lpm

file_refer = lpm.refer("Program Specification.docx")
output_file = lpm.input("Filled-Access Control.xlsx")


def fill_data_to_access_control_excel(inc_data):
    wb = openpyxl.Workbook()

    print("fill in excel")
    count = 0

    for item in inc_data:
        api_id = item.get("API ID", "Unknown")
        print(f"fill {api_id}")
        sheet = wb.create_sheet(api_id)

        sheet.append(["Required JWT Access Role", "Required UAM Data Permission (Access Rights)"])

        bold_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = bold_font

        required_jwt = item.get("Required JWT Access Role", "")
        required_uam = item.get("Required UAM Data Permission (Access Rights)", "")
        sheet.append([required_jwt, required_uam])
        count += 1

    readme_ws = wb.create_sheet("README", 0)
    readme_ws.append(["API ID"])

    for item in inc_data:
        api_id = item.get("API ID", "Unknown")
        readme_ws.append([api_id])

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_file)
    print(f"finish {count}")


def extract_access_control_table():
    doc = Document(file_refer)

    api_sections = []
    api_section_data = {}
    current_section = None
    table_count = len(doc.tables)  # Note
    # print(table_count)

    for i, element in enumerate(doc.element.body):
        if isinstance(element, docx.oxml.text.paragraph.CT_P):
            paragraph = docx.text.paragraph.Paragraph(element, doc)
            text = paragraph.text.strip()

            if text.startswith("AA-") or text.startswith("BB-"):
                if current_section:
                    api_sections.append(api_section_data)
                    api_section_data = {}

                current_section = text
                api_section_data["API Section"] = current_section
                print("----------------------------------------------------------")
                print(current_section)
            elif text.startswith("API ID"):
                strip = text.split(":")[-1].strip()
                if strip not in current_section:
                    print("not matching API ID")
                    api_section_data["API ID"] = current_section[:12]  # handle prog spec API ID typo
                else:
                    api_section_data["API ID"] = text.split(":")[-1].strip()

                print(text.split(":")[-1].strip())
        elif isinstance(element, docx.oxml.table.CT_Tbl):
            table = docx.table.Table(element, doc)
            # print("ping table")
            # print(len(table.rows))

            first_row = table.rows[0]
            if len(first_row.cells) > 1:
                first_cell = first_row.cells[0]
                second_cell = first_row.cells[1]
                first_cell_text = first_cell.text.strip()
                second_cell_text = second_cell.text.strip()

                if first_cell_text == "Required JWT Access Role" \
                        and second_cell_text == "Required UAM Data Permission (Access Rights)":

                    for row in table.rows[1:]:
                        role_value = row.cells[0].text.strip()
                        rights_value = row.cells[1].text.strip()

                        api_section_data["Required JWT Access Role"] = role_value
                        api_section_data["Required UAM Data Permission (Access Rights)"] = rights_value

                        print(f"   Required JWT Access Role: {role_value}")
                        print(f"   Required UAM Data Permission (Access Rights): {rights_value}")
        else:
            continue

    if current_section:
        api_sections.append(api_section_data)

    # print(api_sections)
    return api_sections


if __name__ == '__main__':
    data = extract_access_control_table()
    fill_data_to_access_control_excel(data)
