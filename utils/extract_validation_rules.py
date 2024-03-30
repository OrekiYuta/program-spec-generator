import docx
import openpyxl
from docx import Document
from openpyxl.styles import Font

from utils.PathManager import load_path_manager as lpm

file_refer = lpm.refer("Program Specification.docx")
data_validation_refer = lpm.refer(
    "Data Validation.xlsx")
output_file = lpm.input("Filled-Validation Rules.xlsx")


def fill_data_to_validation_rules_excel(inc_data):
    wb = openpyxl.Workbook()

    print("fill in excel")
    count = 0

    for item in inc_data:
        api_id = item.get("API ID", "Unknown")
        print(f"fill {api_id}")
        sheet = wb.create_sheet(api_id)

        sheet.append(["Interface", "Entity name", "Field name", "Field type", "Validation Rules"])

        bold_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = bold_font

        if "VR" in item:
            vr_data = item["VR"]
            for vr_item in vr_data:
                interface = vr_item.get("Interface", "")
                entity_name = vr_item.get("Entity name", "")
                field_name = vr_item.get("Field name", "")
                field_type = vr_item.get("Field type", "")
                validation_rules = vr_item.get("Validation Rules", "")
                sheet.append([interface, entity_name, field_name, field_type, validation_rules])

        count += 1

    readme_ws = wb.create_sheet("README", 0)  # Add README sheet at the beginning
    readme_ws.append(["API ID"])

    for item in inc_data:
        api_id = item.get("API ID", "Unknown")
        readme_ws.append([api_id])

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_file)
    print(f"finish {count}")


def extract_field_validations():
    wb = openpyxl.load_workbook(data_validation_refer)

    field_validations = {}

    for sheet in wb:
        if not sheet.title.startswith("FS"):
            continue

        ws = wb[sheet.title]
        field_validations[sheet.title] = {}

        count = 1
        for row in ws:
            if count < 3:
                count += 1
                continue

            function = row[1].value
            field = row[2].value
            validation_rule_1 = row[27].value
            validation_rule_2 = row[28].value
            validation_rule_3 = row[29].value
            field_validations[sheet.title][function] = {}
            field_validations[sheet.title][function][field] = [i for i in
                                                               [validation_rule_1, validation_rule_2, validation_rule_3]
                                                               if i]

            print(f"Function: {function}")
            print(f"\tField: {field}")
            print(
                f"\t\tValidation Rules: [{', '.join([i for i in [validation_rule_1, validation_rule_2, validation_rule_3] if i])}]")

            count += 1

    wb.close()


def extract_validation_rules_table():
    doc = Document(file_refer)

    api_sections = []
    api_section_data = {}
    current_section = None
    table_count = len(doc.tables)  # Note
    # print(table_count)

    for i, element in enumerate(doc.element.body):
        if isinstance(element, docx.oxml.text.paragraph.CT_P):
            paragraph = docx.text.paragraph.Paragraph(element, doc)
            text = paragraph.text.strip()

            if text.startswith("AA-") or text.startswith("BB-"):
                if current_section:
                    api_sections.append(api_section_data)
                    api_section_data = {}

                current_section = text
                api_section_data["API Section"] = current_section
                print("----------------------------------------------------------")
                print(current_section)
            elif text.startswith("API ID"):
                strip = text.split(":")[-1].strip()
                if strip not in current_section:
                    print("not matching API ID")
                    api_section_data["API ID"] = current_section[:12]  # handle prog spec API ID typo
                else:
                    api_section_data["API ID"] = text.split(":")[-1].strip()

                print(text.split(":")[-1].strip())

        elif isinstance(element, docx.oxml.table.CT_Tbl):
            table = docx.table.Table(element, doc)
            # print("ping table")
            # print(len(table.rows))

            first_row = table.rows[0]
            if len(first_row.cells) > 1:
                first_cell = first_row.cells[0]
                second_cell = first_row.cells[1]
                first_cell_text = first_cell.text.strip()
                second_cell_text = second_cell.text.strip()

                if first_cell_text == "Interface" \
                        and second_cell_text == "Entity name":

                    vr_table = []
                    for row in table.rows[1:]:
                        vr_item = {}
                        i_value = row.cells[0].text.strip()
                        en_value = row.cells[1].text.strip()
                        fn_value = row.cells[2].text.strip()
                        ft_value = row.cells[3].text.strip()
                        vr_value = row.cells[4].text.strip()

                        vr_item["Interface"] = i_value
                        vr_item["Entity name"] = en_value
                        vr_item["Field name"] = fn_value
                        vr_item["Field type"] = ft_value
                        vr_item["Validation Rules"] = vr_value
                        vr_table.append(vr_item)
                        # print(f"   Interface: {i_value}")
                        # print(f"   Entity name: {en_value}")
                        # print(f"   Field type: {fn_value}")
                        # print(f"   Field type: {ft_value}")
                        # print(f"   Validation Rules: {vr_value}")
                    api_section_data["VR"] = vr_table
                    print(vr_table)

        else:
            continue

    if current_section:
        api_sections.append(api_section_data)

    # print(api_sections)

    return api_sections


if __name__ == '__main__':
    # extract_field_validations()
    data = extract_validation_rules_table()
    fill_data_to_validation_rules_excel(data)
